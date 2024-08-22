from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from .models import CustomUser, Table, Column, Row, Project, ActivityLog
from django.db import IntegrityError
from .forms import CustomUserCreationForm, CustomAuthenticationForm, TableForm, ColumnForm, DataForm, ProjectForm
from django.db.models import Prefetch, Count
from django.urls import reverse
from django.contrib import messages
from .forms import CustomUserChangeForm
import csv
import openpyxl
from openpyxl.utils import get_column_letter
import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile
from io import TextIOWrapper
import json
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.shortcuts import redirect
from django.urls import reverse
from .models import CustomUser
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponseForbidden
from django.db.models import Q
from .models import ActivityLog

from django.contrib.auth import get_user_model

import requests
import base64
from django.http import HttpResponse
from django.urls import reverse
import zlib

def generate_mermaid(project):
    tables = project.tables.all()
    mermaid_code = "erDiagram\n"
    
    for table in tables:
        table_name = table.name.replace(" ", "_")
        mermaid_code += f"    {table_name} {{\n"
        for column in table.columns.all():
            column_not_space = column.name.replace(" ", "_")
            mermaid_code += f"        string {column_not_space}\n"
        mermaid_code += "    }\n"
    
    for table in tables:
        for column in table.columns.all():
            if column.related_table:
                mermaid_code += f"    {table.name.replace(" ", "_")} ||--o{{ {column.related_table.name.replace(" ", "_")} : {column.name.replace(" ", "_")}\n"
    
    return mermaid_code

@login_required
def export_erd(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if project.user != request.user and request.user not in project.shared_users.all():
        return HttpResponse("Anda tidak memiliki akses ke proyek ini.", status=403)
    
    mermaid_code = generate_mermaid(project)
    
    context = {
        'mermaid_code': mermaid_code,
    }
    
    return render(request, 'erd.html', context)

def log_activity(user, table, action, description):
    ActivityLog.objects.create(
        user=user,
        table=table,
        action=action,
        description=description
    )
    
@login_required
def print_table(request, table_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    columns = table.columns.all()
    rows = table.rows.all()
    data = []
    
    for row in rows:
        row_data = []
        for column in columns:
            value = row.data_json.get(str(column.id), '')
            if column.related_table and value:
                try:
                    related_row = Row.objects.filter(id=int(value), table=column.related_table).first()
                    if related_row:
                        value = related_row.data_json.get(str(related_row.table.columns.first().id), '')
                except ValueError:
                    value = ''
            row_data.append(value)
        data.append(row_data)
    
    context = {
        'table': table,
        'columns': columns,
        'data': data,
    }
    return render(request, 'print_table.html', context)

def index(request):
    return redirect('dashboard')

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Akun berhasil dibuat. Selamat datang!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = CustomAuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, 'Anda berhasil masuk.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, 'Anda telah keluar dari sistem.')
    return redirect('login')

@login_required
def dashboard_view(request):
    tables = Table.objects.filter(user=request.user, project__isnull=True).annotate(row_count=Count('rows'))
    return render(request, 'dashboard.html', {'tables': tables})

@login_required
def create_table_view(request):
    if request.method == 'POST':
        form = TableForm(request.POST)
        if form.is_valid():
            table = form.save(commit=False)
            table.user = request.user
            table.save()
            log_activity(request.user, table, 'create_table', f'Membuat tabel {table.name}')
            messages.success(request, 'Tabel berhasil dibuat.')
            return redirect('table_detail', table_id=table.id)
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = TableForm()
    return render(request, 'dashboard.html', {'form': form})

@login_required
def table_detail_view(request, table_id):
    table = get_object_or_404(Table, id=table_id)
    if table.user != request.user and (not table.project or request.user not in table.project.shared_users.all()):
        return HttpResponseForbidden("Anda tidak memiliki akses ke tabel ini.")
    columns = table.columns.all()
    rows = table.rows.all()

    # Tambahkan logika pencarian
    search_query = request.GET.get('search', '')
    if search_query:
        filtered_rows = []
        for row in rows:
            if any(search_query.lower() in str(value).lower() for value in row.data_json.values()):
                filtered_rows.append(row)
        rows = filtered_rows

    data = []
    for row in rows:
        row_data = {'id': row.id, 'values': []}
        for column in columns:
            value = row.data_json.get(str(column.id), '')
            if column.related_table and value:
                try:
                    related_row = Row.objects.filter(id=int(value), table=column.related_table).first()
                    if related_row:
                        display_value = related_row.data_json.get(str(related_row.table.columns.first().id), '')
                        row_data['values'].append(display_value)
                    else:
                        row_data['values'].append('')
                except ValueError:
                    row_data['values'].append('')
            else:
                row_data['values'].append(value)
        data.append(row_data)
    
    # Tambahkan data untuk kolom dengan relasi
    for column in columns:
        if column.related_table:
            if table.project:
                related_rows = Row.objects.filter(table=column.related_table, table__project=table.project)
            else:
                related_rows = Row.objects.filter(table=column.related_table)
            column.related_table_rows = [
                {
                    'id': row.id,
                    'display_value': row.data_json.get(str(row.table.columns.first().id), '')
                }
                for row in related_rows
            ]
    
    if table.project:
        user_tables = Table.objects.filter(user=request.user, project=table.project).exclude(id=table_id)
    else:
        user_tables = Table.objects.filter(user=request.user, project__isnull=True).exclude(id=table_id)
    
    api_url = request.build_absolute_uri(reverse('api_table-detail', args=[table_id]))

    # Paginasi
    items_per_page = request.GET.get('items_per_page', '15')
    if items_per_page == 'all':
        page_obj = data
    else:
        items_per_page = int(items_per_page)
        paginator = Paginator(data, items_per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    # Ganti 'data' dengan 'page_obj' dalam konteks

    # Tambahkan ini untuk menangani permintaan AJAX untuk edit dan hapus
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.method == 'POST':
            action = request.POST.get('action')
            row_id = request.POST.get('row_id')
            if action == 'edit':
                row = get_object_or_404(Row, id=row_id, table=table)
                data_json = {}
                for key, value in request.POST.items():
                    if key.startswith('column_'):
                        column_id = key.split('_')[1]
                        column = get_object_or_404(Column, id=column_id, table=table)
                        data_json[column_id] = value
                row.data_json = data_json
                row.save()
                log_activity(request.user, table, 'edit', f'Mengedit data dengan ID {row_id}')
                messages.success(request, 'Data berhasil diubah.')
                return JsonResponse({'success': True, 'message': 'Data berhasil diubah'})
            elif action == 'delete':
                row = get_object_or_404(Row, id=row_id, table=table)
                row.delete()
                log_activity(request.user, table, 'delete', f'Menghapus data dengan ID {row_id}')
                messages.success(request, 'Data berhasil dihapus.')
                return JsonResponse({'success': True, 'message': 'Data berhasil dihapus'})
            elif action == 'bulk_delete':
                row_ids = json.loads(request.POST.get('row_ids', '[]'))
                try:
                    # Lakukan penghapusan data berdasarkan row_ids
                    Row.objects.filter(id__in=row_ids).delete()
                    log_activity(request.user, table, 'bulk_delete', f'Menghapus {len(row_ids)} data')
                    messages.success(request, 'Data berhasil dihapus.')
                    return JsonResponse({'success': True, 'message': 'Data berhasil dihapus.'})
                except Exception as e:
                    messages.error(request, f'Terjadi kesalahan: {str(e)}')
                    return JsonResponse({'success': False, 'message': str(e)})
            elif action == 'add':
                data_json = {}
                for key, value in request.POST.items():
                    if key.startswith('column_'):
                        column_id = key.split('_')[1]
                        column = get_object_or_404(Column, id=column_id, table=table)
                        if column.related_table:
                            data_json[column_id] = value
                        else:
                            data_json[column_id] = value
                row = Row.objects.create(table=table, data_json=data_json)
                log_activity(request.user, table, 'add', 'Menambahkan data baru')
                messages.success(request, 'Data berhasil ditambahkan.')
                return JsonResponse({'success': True, 'message': 'Data berhasil ditambahkan'})
        return JsonResponse({'success': False, 'message': 'Permintaan tidak valid'})

    # Ambil 3 aktivitas terakhir terkait tabel ini
    recent_activities = ActivityLog.objects.filter(
        Q(table=table) | Q(row__table=table)
    ).order_by('-timestamp')[:3]

    return render(request, 'table_detail.html', {
        'table': table,
        'columns': columns,
        'page_obj': page_obj,
        'user_tables': user_tables,
        'api_url': api_url,
        'total_rows': table.rows.count(),
        'items_per_page': items_per_page,
        'recent_activities': recent_activities,
    })

@login_required
def add_column_view(request, table_id):
    table = get_object_or_404(Table, id=table_id)
    if table.user != request.user and (not table.project or request.user not in table.project.shared_users.all()):
        return HttpResponseForbidden("Anda tidak memiliki akses untuk menambah kolom pada tabel ini.")
    if request.method == 'POST':
        form = ColumnForm(request.POST, user=request.user, current_table=table)
        if form.is_valid():
            column = form.save(commit=False)
            column.table = table
            column.save()
            log_activity(request.user, table, 'add_column', f'Menambahkan kolom {column.name}')
            messages.success(request, 'Kolom berhasil ditambahkan.')
            return redirect('table_detail', table_id=table.id)
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = ColumnForm(user=request.user, current_table=table)
    
    user_tables = Table.objects.filter(Q(user=request.user) | Q(project__shared_users=request.user)).exclude(id=table_id)
    
    return render(request, 'table_detail.html', {
        'form': form, 
        'table': table,
        'user_tables': user_tables
    })

@login_required
def add_data_view(request, table_id):
    table = get_object_or_404(Table, id=table_id)
    if table.user != request.user and (not table.project or request.user not in table.project.shared_users.all()):
        return HttpResponseForbidden("Anda tidak memiliki akses untuk menambah data pada tabel ini.")
    columns = table.column_set.all()
    
    if request.method == 'POST':
        form = DataForm(request.POST, request.FILES, table=table)
        if form.is_valid():
            row = Row.objects.create(table=table)
            data_json = {}
            for field_name, value in form.cleaned_data.items():
                if field_name.startswith('column_'):
                    column_id = field_name.split('_')[1]
                    column = Column.objects.get(id=column_id)
                    if column.related_table:
                        data_json[column_id] = str(value.id) if value else ''
                    else:
                        data_json[column_id] = str(value) if value is not None else ''
            row.data_json = data_json
            row.save()
            log_activity(request.user, table, 'add_data', 'Menambahkan data baru')
            messages.success(request, 'Data berhasil ditambahkan.')
            return redirect('table_detail', table_id=table.id)
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = DataForm(table=table)
    return render(request, 'modal.html', {'form': form, 'table': table, 'columns': columns})

@login_required
def edit_data_view(request, row_id):
    row = get_object_or_404(Row, id=row_id)
    table = row.table
    if table.user != request.user and (not table.project or request.user not in table.project.shared_users.all()):
        return HttpResponseForbidden("Anda tidak memiliki akses untuk mengedit data pada tabel ini.")
    columns = table.columns.all()

    if request.method == 'POST':
        data_json = {}
        for column in columns:
            value = request.POST.get(f'column_{column.id}')
            if column.related_table:
                data_json[column.name] = value if value else None
            else:
                data_json[column.name] = value
        row.data_json = data_json
        row.save()
        log_activity(request.user, table, 'edit_data', f'Mengedit data dengan ID {row_id}')
        messages.success(request, 'Data berhasil diubah.')
        return redirect('table_detail', table_id=table.id)

    return render(request, 'table_detail.html', {'row': row, 'columns': columns})

@login_required
def delete_data_view(request, row_id):
    row = get_object_or_404(Row, id=row_id)
    table = row.table
    if table.user != request.user and (not table.project or request.user not in table.project.shared_users.all()):
        return HttpResponseForbidden("Anda tidak memiliki akses untuk menghapus data pada tabel ini.")
    table_id = table.id
    row.delete()
    log_activity(request.user, table, 'delete_data', f'Menghapus data dengan ID {row_id}')
    messages.success(request, 'Data berhasil dihapus.')
    return redirect('table_detail', table_id=table_id)

@login_required
def edit_table_view(request, table_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    if request.method == 'POST':
        form = TableForm(request.POST, instance=table)
        if form.is_valid():
            form.save()
            log_activity(request.user, table, 'edit_table', f'Mengedit tabel {table.name}')
            messages.success(request, 'Tabel berhasil diubah.')
            return redirect('table_detail', table_id=table.id)
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    return render(request, 'table_detail.html', {'form': form, 'table': table})

@login_required
def delete_table_view(request, table_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    table_name = table.name
    if table.project:
        project = table.project
        table.delete()
        log_activity(request.user, None, 'delete_table', f'Menghapus tabel {table_name} dari proyek {project.name}')
        messages.success(request, 'Tabel berhasil dihapus.')
        return redirect('project_detail', project_id=project.id)
    else:
        table.delete()
        log_activity(request.user, None, 'delete_table', f'Menghapus tabel {table_name}')
        messages.success(request, 'Tabel berhasil dihapus.')
    return redirect('dashboard')

@login_required
def get_columns(request, table_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    columns = Column.objects.filter(table=table)
    data = [{'id': column.id, 'name': column.name} for column in columns]
    return JsonResponse({'columns': data})
    
@login_required
def profile_view(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profil berhasil diperbarui.')
            return redirect('profile')
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
    else:
        form = CustomUserChangeForm(instance=request.user)
    
    return render(request, 'profile.html', {'form': form})

@login_required
def export_table_data(request, table_id, format):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    columns = table.columns.all()
    rows = table.rows.all()

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{table.name}.csv"'
        writer = csv.writer(response)

        # Tulis header
        header = [column.name for column in columns]
        writer.writerow(header)

        # Tulis data
        for row in rows:
            row_data = []
            for column in columns:
                value = row.data_json.get(str(column.id), '')
                if column.related_table and value:
                    related_row = Row.objects.filter(id=int(value), table=column.related_table).first()
                    if related_row:
                        value = related_row.data_json.get(str(related_row.table.columns.first().id), '')
                row_data.append(value)
            writer.writerow(row_data)

        messages.success(request, 'Data berhasil diekspor ke CSV.')
        return response

    elif format == 'excel':
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = table.name

        # Tulis header
        for col_num, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column.name

        # Tulis data
        for row_num, row in enumerate(rows, 2):
            for col_num, column in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = row.data_json.get(str(column.id), '')
                if column.related_table and value:
                    related_row = Row.objects.filter(id=int(value), table=column.related_table).first()
                    if related_row:
                        value = related_row.data_json.get(str(related_row.table.columns.first().id), '')
                cell.value = value

        # Atur lebar kolom
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{table.name}.xlsx"'
        workbook.save(response)
        messages.success(request, 'Data berhasil diekspor ke Excel.')
        return response

    messages.error(request, "Format tidak didukung")
    return HttpResponse("Format tidak didukung", status=400)

@login_required
def import_data(request, table_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
        elif file_extension in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            reader = [
                {headers[i]: cell.value for i, cell in enumerate(row)}
                for row in sheet.iter_rows(min_row=2)
            ]
        else:
            messages.error(request, 'Format file tidak didukung. Gunakan CSV atau Excel.')
            return redirect('table_detail', table_id=table.id)

        columns = {column.name: column for column in table.columns.all()}
        for row in reader:
            new_row = Row(table=table)
            data_json = {}
            for column_name, value in row.items():
                if column_name in columns:
                    column = columns[column_name]
                    if column.related_table:
                        # Cari data yang cocok di tabel terkait
                        related_rows = Row.objects.filter(table=column.related_table)
                        related_column = column.related_table.columns.first()
                        if related_column:
                            for related_row in related_rows:
                                if related_row.data_json.get(str(related_column.id)) == str(value):
                                    data_json[str(column.id)] = str(related_row.id)
                                    break
                    else:
                        data_json[str(column.id)] = str(value) if value is not None else ''
            new_row.data_json = data_json
            new_row.save()

        log_activity(request.user, table, 'import_data', 'Mengimpor data')
        messages.success(request, 'Data berhasil diimpor.')
        return redirect('table_detail', table_id=table.id)

    return render(request, 'import_data.html', {'table': table})

@login_required
def import_table(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            headers = reader.fieldnames
        elif file_extension in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            reader = [
                {headers[i]: cell.value for i, cell in enumerate(row)}
                for row in sheet.iter_rows(min_row=2)
            ]
        else:
            messages.error(request, 'Format file tidak didukung. Gunakan CSV atau Excel.')
            return redirect('dashboard')

        table_name = request.POST.get('table_name', 'Tabel Impor')
        new_table = Table.objects.create(name=table_name, user=request.user)

        for header in headers:
            Column.objects.create(name=header, table=new_table)

        for row in reader:
            new_row = Row(table=new_table)
            data_json = {
                str(new_table.columns.get(name=column_name).id): str(value) if value is not None else ''
                for column_name, value in row.items()
            }
            new_row.data_json = data_json
            new_row.save()

        log_activity(request.user, new_table, 'import_table', f'Mengimpor tabel {new_table.name}')
        messages.success(request, 'Tabel baru berhasil diimpor.')
        return redirect('table_detail', table_id=new_table.id)

    return render(request, 'import_table.html')

@login_required
def edit_column_view(request, table_id, column_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    column = get_object_or_404(Column, id=column_id, table=table)
    
    if request.method == 'POST':
        form = ColumnForm(request.POST, instance=column, user=request.user, current_table=table)
        if form.is_valid():
            form.save()
            log_activity(request.user, table, 'edit_column', f'Mengedit kolom {column.name}')
            messages.success(request, 'Kolom berhasil diubah.')
            return redirect('table_detail', table_id=table.id)
        else:
            messages.error(request, 'Terjadi kesalahan saat mengubah kolom. Silakan coba lagi.')
    else:
        form = ColumnForm(instance=column, user=request.user, current_table=table)
    
    return render(request, 'table_detail.html', {'form': form, 'table': table, 'column': column})

@login_required
def delete_column_view(request, table_id, column_id):
    table = get_object_or_404(Table, id=table_id, user=request.user)
    column = get_object_or_404(Column, id=column_id, table=table)
    
    if request.method == 'POST':
        column.delete()
        log_activity(request.user, table, 'delete_column', f'Menghapus kolom {column.name}')
        messages.success(request, 'Kolom berhasil dihapus.')
        return redirect('table_detail', table_id=table.id)
    
    messages.warning(request, 'Konfirmasi penghapusan kolom.')
    return render(request, 'table_detail.html', {'table': table, 'column': column})

@login_required
def project_list(request):
    owned_projects = Project.objects.filter(user=request.user)
    shared_projects = request.user.shared_projects.all()
    projects = list(owned_projects) + list(shared_projects)
    return render(request, 'project_list.html', {'projects': projects})

@login_required
def create_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.user = request.user
            project.save()
            log_activity(request.user, None, 'create_project', f'Membuat proyek {project.name}')
            messages.success(request, 'Proyek berhasil dibuat.')
            return JsonResponse({
                'success': True,
                'message': 'Proyek berhasil dibuat.',
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Terjadi kesalahan. Silakan periksa form Anda.',
                'errors': form.errors
            }, status=400)
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan.'}, status=405)

@login_required
def project_detail(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if project.user != request.user and request.user not in project.shared_users.all():
        return JsonResponse({'success': False, 'message': 'Anda tidak memiliki akses ke proyek ini.'}, status=403)
    tables = Table.objects.filter(project=project)
    return render(request, 'project_detail.html', {'project': project, 'tables': tables})

@login_required
def create_table(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if project.user != request.user and request.user not in project.shared_users.all():
        return JsonResponse({'success': False, 'message': 'Anda tidak memiliki akses ke proyek ini.'}, status=403)
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            form = TableForm(request.POST)
            if form.is_valid():
                table = form.save(commit=False)
                table.user = project.user  # Set user to project owner
                table.project = project
                table.save()
                log_activity(request.user, table, 'create_table', f'Membuat tabel {table.name} di proyek {project.name}')
                messages.success(request, 'Tabel berhasil dibuat.')
                return JsonResponse({'success': True, 'redirect_url': reverse('project_detail', args=[project_id])})
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                return JsonResponse({'success': False, 'redirect_url': reverse('project_detail', args=[project_id])}, status=400)
        except Exception as e:
            messages.error(request, str(e))
            return JsonResponse({'success': False}, status=400)
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan.'}, status=405)

@login_required
def delete_project(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    if request.method == 'POST':
        project.delete()
        log_activity(request.user, None, 'delete_project', f'Menghapus proyek {project.name}')
        messages.success(request, 'Proyek berhasil dihapus.')
        return redirect('project_list')
    return redirect('project_detail', project_id=project_id)

@login_required
def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if project.user != request.user and request.user not in project.shared_users.all():
        return JsonResponse({'success': False, 'message': 'Anda tidak memiliki akses ke proyek ini.'}, status=403)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            log_activity(request.user, None, 'edit_project', f'Mengedit proyek {project.name}')
            messages.success(request, 'Proyek berhasil diperbarui.')
            return JsonResponse({'success': True, 'message': 'Proyek berhasil diperbarui.', 'redirect_url': reverse('project_detail', args=[project_id])})
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form Anda.')
            return JsonResponse({'success': False, 'message': 'Terjadi kesalahan. Silakan periksa form Anda.'}, status=400)
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan.'}, status=405)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_users(request):
    term = request.GET.get('term', '')
    users = CustomUser.objects.filter(
        Q(username__icontains=term) | Q(email__icontains=term)
    ).exclude(id=request.user.id).exclude(is_superuser=True)[:10]  # Batasi hasil pencarian dan kecualikan superuser
    data = [
        {
            'id': user.id,
            'username': user.username,
            'avatar': user.avatar.url if user.avatar else None,
            'match_index': user.username.lower().index(term.lower()) if term.lower() in user.username.lower() else float('inf')
        } 
        for user in users
    ]
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def shared_users(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    shared_users = project.shared_users.all()
    data = [{'id': user.id, 'username': user.username, 'avatar': user.avatar.url if user.avatar else None} for user in shared_users]
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def share_project(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    user_ids = request.data.get('users', [])
    users_to_share = CustomUser.objects.filter(id__in=user_ids)
    project.shared_users.set(users_to_share)
    log_activity(request.user, None, 'share_project', f'Membagikan proyek {project.name} ke {", ".join([user.username for user in users_to_share])}')
    messages.success(request, 'Proyek berhasil dibagikan.')
    return Response({'success': True, 'message': 'Proyek berhasil dibagikan.'})

@login_required
def import_table_to_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if project.user != request.user and request.user not in project.shared_users.all():
        return JsonResponse({'success': False, 'message': 'Anda tidak memiliki akses ke proyek ini.'}, status=403)
    
    if request.method == 'POST' and request.FILES.get('file'):
        table_name = request.POST.get('table_name')
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            headers = reader.fieldnames
        elif file_extension in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            reader = [
                {headers[i]: cell.value for i, cell in enumerate(row)}
                for row in sheet.iter_rows(min_row=2)
            ]
        else:
            messages.error(request, 'Format file tidak didukung. Gunakan CSV atau Excel.')
            return redirect('project_detail', project_id=project.id)

        new_table = Table.objects.create(name=table_name, user=request.user, project=project)

        for header in headers:
            Column.objects.create(name=header, table=new_table)

        for row in reader:
            new_row = Row(table=new_table)
            data_json = {
                str(new_table.columns.get(name=column_name).id): str(value) if value is not None else ''
                for column_name, value in row.items()
            }
            new_row.data_json = data_json
            new_row.save()

        log_activity(request.user, new_table, 'import_table', f'Mengimpor tabel {new_table.name} ke proyek {project.name}')
        messages.success(request, 'Tabel baru berhasil diimpor ke dalam proyek.')
        return redirect('project_detail', project_id=project.id)

    messages.error(request, 'Terjadi kesalahan saat mengimpor tabel.')
    return redirect('project_detail', project_id=project.id)